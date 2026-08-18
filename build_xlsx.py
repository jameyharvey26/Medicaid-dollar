from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ARIAL="Arial"
BLUE=Font(name=ARIAL,color="0000FF"); BLACK=Font(name=ARIAL,color="000000"); BOLD=Font(name=ARIAL,bold=True)
TITLE=Font(name=ARIAL,bold=True,size=14); SUB=Font(name=ARIAL,italic=True,color="595959")
SECT=Font(name=ARIAL,bold=True,color="FFFFFF"); HDR=Font(name=ARIAL,bold=True)
YEL=PatternFill("solid",fgColor="FFFF00"); SECTFILL=PatternFill("solid",fgColor="1F4E5F"); GREYFILL=PatternFill("solid",fgColor="F1EFE8")
USD='$#,##0;($#,##0);-'; USD2='$#,##0.00;($#,##0.00);-'; PCT='0.0%'; PCT2='0.00%'

# ============ CANONICAL V.4 LEDGER (per $100 of TOTAL outlays) ============
TOTAL=957400.0; BEN=908839.0; FEDSH=0.647
admin=round((TOTAL-BEN)/TOTAL*100,2)            # 5.07
medicare=round(27774/TOTAL*100,2)               # 2.90
benefits100=round(BEN/TOTAL*100,2)              # 94.93
disb=round(benefits100-medicare,2)              # 92.03
fed=round(FEDSH*100,2); state=round(100-fed,2)  # 64.70 / 35.30
# lanes (collections folded proportionally), per $100 total
cap=496097.0; ffs_g=400002.0; dualcap=106000.0; mcocap=cap-dualcap; coll=15201.0
netden=cap+ffs_g
mco=round(disb*mcocap/netden,2); dual=round(disb*dualcap/netden,2); ffs=round(disb*ffs_g/netden,2)
MLR=0.89; ADMINPCT=0.09
mco_ret=round(mco*(1-MLR),2); dual_ret=round(dual*(1-MLR),2); plan_ret=round(mco_ret+dual_ret,2)
earnings=0.76; adm_marg=round(plan_ret-earnings,2)
mco_care=round(mco*MLR,2); dual_care=round(dual*MLR,2); capcare=round(mco_care+dual_care,2)
provin=round(ffs+mco_care+dual_care,2)          # 86.42
fraud=round(1400/TOTAL*100,2)                   # 0.15
# base-5 nodes (FQHC merged into Physicians & clinics)
cms={"Long-term care":194580,"Hospitals":91443,"Other":61284,"Physicians & clinics":32489,"Rx drugs":20206}
csum=sum(cms.values())
hma={"Long-term care":.197,"Hospitals":.238,"Other":.235,"Physicians & clinics":.251,"Rx drugs":.079}
base={k: round(ffs*cms[k]/csum + capcare*hma[k],2) for k in cms}
# behavioral-health cross-cutting carve
carve={"Long-term care":0.38,"Hospitals":1.52,"Other":3.62,"Physicians & clinics":2.09,"Rx drugs":1.90}
bh=round(sum(carve.values()),2)                 # 10.40
final={}
for k in ["Long-term care","Hospitals","Other","Physicians & clinics","Rx drugs"]: final[k]=round(base[k]-carve.get(k,0),2)
final["Behavioral health"]=bh
order=sorted(final,key=lambda k:-final[k])
# per-lane composition of final 6 nodes
ffs_n={k: ffs*cms[k]/csum for k in cms}; mco_n={k: mco_care*hma[k] for k in cms}; dual_n={k: dual_care*hma[k] for k in cms}
BHf=BHm=BHd=0.0
for d,amt in carve.items():
    tot=ffs_n[d]+mco_n[d]+dual_n[d]; fr=amt/tot
    BHf+=ffs_n[d]*fr; BHm+=mco_n[d]*fr; BHd+=dual_n[d]*fr
    ffs_n[d]*=(1-fr); mco_n[d]*=(1-fr); dual_n[d]*=(1-fr)
ffs_n["Behavioral health"]=BHf; mco_n["Behavioral health"]=BHm; dual_n["Behavioral health"]=BHd
# IPF beneficiary cross-tab
G=["Children","Adults","Disabled","Aged"]; gsh=[0.156,0.342,0.289,0.213]; gt={g:round(provin*s,2) for g,s in zip(G,gsh)}
seed={"Long-term care":[.02,.10,1.0,1.0],"Hospitals":[.85,1.0,.75,.45],"Other":[.95,.95,.80,.55],
 "Physicians & clinics":[1.05,1.0,.60,.40],"Behavioral health":[.55,1.05,1.35,.30],"Rx drugs":[.40,.85,1.05,.65]}
M={p:{g:seed[p][i] for i,g in enumerate(G)} for p in order}
for _ in range(100):
    for p in order:
        s=sum(M[p].values())
        for g in G: M[p][g]*=final[p]/s
    for g in G:
        s=sum(M[p][g] for p in order)
        for p in order: M[p][g]*=gt[g]/s
pie={p:{g:M[p][g]/sum(M[p].values()) for g in G} for p in order}

wb=Workbook()
def put(ws,cell,val,font=None,fmt=None,fill=None,align=None,wrap=False):
    c=ws[cell]; c.value=val; c.font=font or BLACK
    if fmt: c.number_format=fmt
    if fill: c.fill=fill
    if align or wrap: c.alignment=Alignment(horizontal=align,wrap_text=wrap,vertical="top")
    return c
def section(ws,cell,text,cols):
    put(ws,cell,text,font=SECT,fill=SECTFILL); row=cell[1:]
    for col in cols: ws[f'{col}{row}'].fill=SECTFILL

# ---------- READ ME ----------
ws=wb.active; ws.title="Read me"
ws.column_dimensions['A'].width=4; ws.column_dimensions['B'].width=106
put(ws,'B2',"$100 Medicaid Dollars - Data, Sources & Assumptions",TITLE)
put(ws,'B3',"Companion workbook to the $100 Medicaid Dollars infographic (DRAFT V.4 - Public Comment). Vintage: FY2024 (federal fiscal year) unless noted.",SUB)
put(ws,'B5',"What this is",BOLD)
put(ws,'B6',"A transparent build of every figure behind the infographic. The unit is $100 of total Medicaid spending - 'the Medicaid "
            "dollar' - formed by combining the federal share ($64.70) and the state share ($35.30). It is run through the state agency "
            "(administration and Medicare-premium losses peel off), disbursed across three payer lanes kept separate to the provider phase "
            "(non-dual MCO, dual MCO, fee-for-service), split into six provider nodes, and finally read as the share of each service consumed "
            "by each beneficiary group. Every input is sourced; every modeling choice is logged.",BLACK,wrap=True)
ws.row_dimensions[6].height=74
put(ws,'B8',"Color & format conventions",BOLD)
put(ws,'B9',"Blue text = hard-coded input taken directly from a published source (see Sources tab).",BLUE)
put(ws,'B10',"Black text = value derived inside the workbook.",BLACK)
put(ws,'B11',"Yellow fill = key assumption or known data gap that needs attention (see Assumptions tab).",BLACK,fill=YEL)
put(ws,'B13',"Tabs",BOLD)
for i,(t,d) in enumerate([
 ("Flow model","The $100 combined dollar; sequential administration and Medicare-premium peels; the three-lane payer split; the dual-capitation derivation; FFS categories and the HMA managed-care key; the six-node provider phase (FQHC merged into Physicians & clinics; behavioral health carved cross-cutting)."),
 ("Beneficiaries","Eligibility-group spending shares (FY2023) mapped to four groups, and the IPF-balanced service-by-group cross-tab behind the six pies."),
 ("Leakages","Plan-retained dollars off each capitated lane (admin / margin / public-company earnings) and documented fraud."),
 ("Sources","Full citation and URL for every figure."),
 ("Assumptions","Every modeling decision, mapping choice, and data gap, numbered A1-A12."),
]):
    put(ws,f'B{14+i}',f"- {t} - {d}",BLACK,wrap=True); ws.row_dimensions[14+i].height=44
put(ws,'B20',"Headline reconciliation (FY2024, per $100 total)",BOLD)
for i,t in enumerate([
 "Total Medicaid spending (benefits + admin): $957.4B  |  benefits only: $908.8B.",
 "Federal share 64.7%, state 35.3% -> the Medicaid dollar = $100 combined = federal $64.70 + state $35.30.",
 "State agency peels: administration $5.07, then Medicare premiums for duals $2.90  ->  $92.03 disbursed.",
 "Three payer lanes (per $100 total): MCO $40.06, dual managed care $10.89, fee-for-service $41.08 (sum $92.03).",
 "Plan retention $5.61 peels off the capitated lanes  ->  $86.42 reaches providers; documented fraud $0.15  ->  $86.27 delivered.",
 "NHE (calendar 2024) reports Medicaid at $931.7B - a different basis; used only as a cross-check.",
]):
    put(ws,f'B{21+i}',f"- {t}",BLACK,wrap=True); ws.row_dimensions[21+i].height=26

# ---------- FLOW MODEL ----------
fm=wb.create_sheet("Flow model")
for col,w in [('A',58),('B',16),('C',12),('D',18),('E',42)]:
    fm.column_dimensions[col].width=w
put(fm,'A1',"$100 Medicaid Dollars - Flow Model - FY2024",TITLE)
put(fm,'A2',"Per $100 of total Medicaid outlays (the combined federal + state dollar). $ in millions unless the column says per $100.",SUB)
section(fm,'A4',"ANCHORS",['A','B','C','D','E'])
put(fm,'A5',"Total Medicaid spending (benefits + administration)"); put(fm,'B5',957400,BLUE,USD); put(fm,'C5',"$M"); put(fm,'D5',"MACStats Feb 2026, Ex.16")
put(fm,'A6',"Total Medicaid benefit spending"); put(fm,'B6',908839,BLUE,USD); put(fm,'C6',"$M"); put(fm,'D6',"MACStats Ex.17")
put(fm,'A7',"Implied administration (total - benefits)"); put(fm,'B7','=B5-B6',BLACK,USD); put(fm,'C7',"$M"); put(fm,'D7',"derived")
put(fm,'A8',"Federal share of benefit spending"); put(fm,'B8',0.647,BLUE,PCT); put(fm,'D8',"MACStats Feb 2026")
put(fm,'A9',"State (non-federal) share"); put(fm,'B9','=1-B8',BLACK,PCT); put(fm,'D9',"derived")
put(fm,'A10',"NHE Medicaid, calendar 2024 (cross-check, diff. basis)"); put(fm,'B10',931700,BLUE,USD); put(fm,'C10',"$M"); put(fm,'D10',"CMS NHE 2024")

section(fm,'A12',"THE MEDICAID DOLLAR ($100 = federal + state combined) AND STATE-AGENCY PEELS",['A','B','C','D','E'])
for cell,t in [('A13','Step'),('B13','Per $100'),('C13',''),('D13','Basis'),('E13','Note')]:
    put(fm,cell,t,HDR,fill=GREYFILL)
put(fm,'A14',"Federal share of the dollar"); put(fm,'B14','=B8*100',BLACK,USD2); put(fm,'D14',"enters from above")
put(fm,'A15',"State share of the dollar"); put(fm,'B15','=B9*100',BLACK,USD2); put(fm,'D15',"merges in")
put(fm,'A16',"The Medicaid dollar (combined)"); put(fm,'B16','=B14+B15',BOLD,USD2)
put(fm,'A17',"less Administration"); put(fm,'B17',-admin,BLACK,USD2); put(fm,'D17','=B7/B5',BLACK,PCT); put(fm,'E17',"admin / total outlays",BLACK,wrap=True)
put(fm,'A18',"less Medicare premiums for duals"); put(fm,'B18',-medicare,BLACK,USD2); put(fm,'D18',27774,BLUE,USD); put(fm,'E18',"$M; Medicaid -> Medicare, peels out",BLACK,wrap=True)
put(fm,'A19',"Disbursable benefits (to the three lanes)"); put(fm,'B19','=B16+B17+B18',BOLD,USD2); put(fm,'E19',"should equal lane sum below",BLACK,wrap=True)

section(fm,'A21',"PAYER-CHANNEL SPLIT - THREE LANES (per $100 total; collections folded proportionally)",['A','B','C','D','E'])
for cell,t in [('A22','Lane'),('B22','Spending ($M)'),('C22','% of net'),('D22','Per $100 total'),('E22','Source / note')]:
    put(fm,cell,t,HDR,fill=GREYFILL)
put(fm,'A23',"Total capitation (single CMS-64 line)"); put(fm,'B23',496097,BLUE,USD); put(fm,'C23','=B23/B27',BLACK,PCT); put(fm,'D23','=C23*$B$19',BLACK,USD2); put(fm,'E23',"Ex.17",BLACK,wrap=True)
put(fm,'A24',"  Lane 1 - MCO (non-dual capitation), MODELED"); put(fm,'B24','=B23-B25',BLACK,USD); put(fm,'C24','=B24/B27',BLACK,PCT); put(fm,'D24','=C24*$B$19',BLACK,USD2); put(fm,'E24',"total capitation minus dual capitation",BLACK,wrap=True)
put(fm,'A25',"  Lane 2 - Dual MCO capitation, MODELED"); put(fm,'B25',106000,BLACK,USD,fill=YEL); put(fm,'C25','=B25/B27',BLACK,PCT); put(fm,'D25','=C25*$B$19',BLACK,USD2); put(fm,'E25',"triangulated (A11); softest number",BLACK,fill=YEL,wrap=True)
put(fm,'A26',"Lane 3 - Fee-for-service, itemized (MEASURED)"); put(fm,'B26',400002,BLUE,USD); put(fm,'C26','=B26/B27',BLACK,PCT); put(fm,'D26','=C26*$B$19',BLACK,USD2); put(fm,'E26',"Ex.17 - sum of categories below",BLACK,wrap=True)
put(fm,'A27',"Net denominator (capitation + FFS)"); put(fm,'B27','=B23+B26',BLACK,USD); put(fm,'E27',"collections folded via this share",BLACK,wrap=True)
put(fm,'A28',"Lane check (per $100 total)"); put(fm,'B28','=D23+D26',BLACK,USD2); put(fm,'D28','=B19',BLACK,USD2); put(fm,'E28',"B28 should equal D28",BLACK,wrap=True)

section(fm,'A30',"PLAN RETENTION CARVE ON CAPITATED LANES (per $100 total)",['A','B','C','D','E'])
put(fm,'A31',"Medical loss ratio (care delivered)"); put(fm,'B31',0.89,BLUE,PCT); put(fm,'E31',"Centene/Molina Medicaid HBR ~88-91%",BLACK,wrap=True)
put(fm,'A32',"MCO care to providers"); put(fm,'B32','=D24*B31',BLACK,USD2)
put(fm,'A33',"Dual care to providers"); put(fm,'B33','=D25*B31',BLACK,USD2)
put(fm,'A34',"Plan retained (both lanes)"); put(fm,'B34','=(D24+D25)*(1-B31)',BLACK,USD2); put(fm,'E34',"plan administration + earnings (Leakages)",BLACK,wrap=True)
put(fm,'A35',"FFS to providers (no carve)"); put(fm,'B35','=D26',BLACK,USD2)
put(fm,'A36',"To providers (FFS + MCO care + dual care)"); put(fm,'B36','=B32+B33+B35',BOLD,USD2)

section(fm,'A38',"FFS CATEGORIES (Ex.17) -> BASE NODES (FQHC merged into Physicians & clinics)",['A','B','C','D','E'])
for cell,t in [('A39','CMS-64 category'),('B39','$M'),('C39','% of FFS'),('D39','Maps to node'),('E39','Note')]:
    put(fm,cell,t,HDR,fill=GREYFILL)
cats=[("Hospital",91443,"Hospitals",""),("Physician",9534,"Physicians & clinics",""),
 ("Other practitioner",3660,"Physicians & clinics","NPs, PAs, therapists"),
 ("Clinic & health center (incl. FQHC)",19295,"Physicians & clinics","FQHC merged here (A4)"),
 ("Dental",6186,"Other",""),("Other acute",55098,"Other","NEMT, DME, EPSDT, hospice, lab (A6)"),
 ("Drugs (net of rebates)",20206,"Rx drugs",""),("Institutional LTSS",64628,"Long-term care",""),
 ("Home- & community-based LTSS",129952,"Long-term care","")]
r=40
for n,a,nd,note in cats:
    put(fm,f'A{r}',n); put(fm,f'B{r}',a,BLUE,USD); put(fm,f'C{r}',f'=B{r}/$B$49',BLACK,PCT); put(fm,f'D{r}',nd); put(fm,f'E{r}',note,BLACK,wrap=True)
    if "FQHC" in note: fm[f'E{r}'].fill=YEL
    r+=1
put(fm,'A49',"FFS subtotal",BOLD); put(fm,'B49','=SUM(B40:B48)',BOLD,USD); put(fm,'C49','=B49/B49',BLACK,PCT)

section(fm,'A51',"HMA T-MSIS MANAGED-CARE KEY (CY2021) -> base nodes",['A','B','C','D','E'])
for cell,t in [('A52','HMA category'),('B52','Share'),('C52','Maps to node'),('D52',''),('E52','Note')]:
    put(fm,cell,t,HDR,fill=GREYFILL)
hmarows=[("Inpatient",0.154,"Hospitals",""),("Outpatient",0.084,"Hospitals",""),
 ("Professional",0.251,"Physicians & clinics","incl. MC FQHC; unbundles hospital MD work (A1)"),
 ("SNF / HCBS (LTSS)",0.197,"Long-term care",""),("Pharmacy",0.079,"Rx drugs",""),
 ("Dental",0.048,"Other",""),("Other + unallocated",0.187,"Other","incl. ~5.5% residual")]
r=53
for n,s,nd,note in hmarows:
    put(fm,f'A{r}',n); put(fm,f'B{r}',s,BLUE,PCT); put(fm,f'C{r}',nd); put(fm,f'E{r}',note,BLACK,wrap=True)
    if "A1" in note: fm[f'E{r}'].fill=YEL
    r+=1
put(fm,'A60',"HMA total",BOLD); put(fm,'B60','=SUM(B53:B59)',BOLD,PCT)

section(fm,'A62',"BEHAVIORAL HEALTH - CROSS-CUTTING CARVE (per $100 total)",['A','B','C','D','E'])
put(fm,'A63',"Behavioral health is 9.3-13% of Medicaid (A12); it is NOT a CMS-64 line. It is carved from the five service lines where it actually sits.",BLACK,wrap=True); fm.merge_cells('A63:E63'); fm.row_dimensions[63].height=28; fm['A63'].fill=YEL
for cell,t in [('A64','Carved from'),('B64','Per $100'),('C64',''),('D64','Setting'),('E64','')]:
    put(fm,cell,t,HDR,fill=GREYFILL)
bhrows=[("Other",carve["Other"],"rehab, residential SUD, community MH, case mgmt"),
 ("Physicians & clinics",carve["Physicians & clinics"],"psychiatry, therapy, CMHC outpatient"),
 ("Rx drugs",carve["Rx drugs"],"psychotropics"),
 ("Hospitals",carve["Hospitals"],"inpatient psych, detox"),
 ("Long-term care",carve["Long-term care"],"psychiatric residential, group homes")]
r=65
for n,a,setn in bhrows:
    put(fm,f'A{r}',n); put(fm,f'B{r}',a,BLACK,USD2,fill=YEL); put(fm,f'D{r}',setn,BLACK,wrap=True); r+=1
put(fm,'A70',"Behavioral health node (total)",BOLD); put(fm,'B70','=SUM(B65:B69)',BOLD,USD2); put(fm,'D70',"~11% of provider dollars (A12)",BLACK,wrap=True)

section(fm,'A71',"PROVIDER PHASE - SIX NODES x THREE LANES (per $100 total, after carves)",['A','B','C','D','E'])
for cell,t in [('A72','Provider node'),('B72','FFS lane'),('C72','MCO lane'),('D72','Dual lane'),('E72','TOTAL')]:
    put(fm,cell,t,HDR,fill=GREYFILL)
r=73
for n in order:
    put(fm,f'A{r}',n); put(fm,f'B{r}',round(ffs_n[n],2),BLACK,USD2); put(fm,f'C{r}',round(mco_n[n],2),BLACK,USD2)
    put(fm,f'D{r}',round(dual_n[n],2),BLACK,USD2); put(fm,f'E{r}',f'=B{r}+C{r}+D{r}',BOLD,USD2); r+=1
put(fm,'A79',"Provider total",BOLD); put(fm,'B79','=SUM(B73:B78)',BOLD,USD2); put(fm,'C79','=SUM(C73:C78)',BOLD,USD2); put(fm,'D79','=SUM(D73:D78)',BOLD,USD2); put(fm,'E79','=SUM(E73:E78)',BOLD,USD2)
put(fm,'A80',"Check vs 'to providers' above"); put(fm,'B80','=B36',BLACK,USD2); put(fm,'E80','=E79',BLACK,USD2); put(fm,'D80',"should match",BLACK)
put(fm,'A81',"less Documented fraud"); put(fm,'B81',-fraud,BLACK,USD2); put(fm,'E81','=E79+B81',BOLD,USD2); put(fm,'D81',"delivered to care",BLACK)

# ---------- BENEFICIARIES ----------
bn=wb.create_sheet("Beneficiaries")
for col,w in [('A',26),('B',12),('C',12),('D',12),('E',12),('F',12),('G',14)]:
    bn.column_dimensions[col].width=w
put(bn,'A1',"Beneficiary Phase - who consumes each service",TITLE)
put(bn,'A2',"Not a money flow. Each service's split is fit by iterative proportional fitting (IPF) to two margins: the service totals (rows) and the group spending totals (columns). Seeded with utilisation patterns. This is the data behind the six pies.",SUB)
bn.merge_cells('A2:G2'); bn.row_dimensions[2].height=42; bn['A2'].alignment=Alignment(wrap_text=True,vertical="top")
section(bn,'A4',"ELIGIBILITY-GROUP SPENDING SHARES (Ex.21, FY2023) -> FOUR GROUPS",['A','B','C','D','E','F','G'])
for cell,t in [('A5','Source group'),('B5','Share'),('C5','Model group'),('D5','Group share'),('E5','Per $100')]:
    put(bn,cell,t,HDR,fill=GREYFILL)
rows=[("Children",0.156,"Children"),("New adult group (expansion)",0.225,"Adults"),("Other adults",0.117,"Adults"),
 ("Individuals with disabilities (<65)",0.289,"Disabled"),("Aged 65+",0.213,"Aged")]
r=6
for n,s,g in rows:
    put(bn,f'A{r}',n); put(bn,f'B{r}',s,BLUE,PCT); put(bn,f'C{r}',g); r+=1
put(bn,'A11',"Children"); put(bn,'B11','=B6',BLACK,PCT); put(bn,'D11','=B6',BLACK,PCT); put(bn,'E11',f'=D11*{provin}',BLACK,USD2)
put(bn,'A12',"Adults (incl. expansion)"); put(bn,'D12','=B7+B8',BLACK,PCT); put(bn,'E12',f'=D12*{provin}',BLACK,USD2)
put(bn,'A13',"Disabled (<65)"); put(bn,'D13','=B9',BLACK,PCT); put(bn,'E13',f'=D13*{provin}',BLACK,USD2)
put(bn,'A14',"Aged (65+)"); put(bn,'D14','=B10',BLACK,PCT); put(bn,'E14',f'=D14*{provin}',BLACK,USD2)
put(bn,'A15',"Total (column margins)",BOLD); put(bn,'D15','=SUM(D11:D14)',BOLD,PCT); put(bn,'E15','=SUM(E11:E14)',BOLD,USD2)

section(bn,'A17',"IPF-BALANCED CROSS-TAB - share of each service consumed by each group",['A','B','C','D','E','F','G'])
for cell,t in [('A18','Service node'),('B18','Children'),('C18','Adults'),('D18','Disabled'),('E18','Aged'),('F18','Row total'),('G18','Node $ (per $100)')]:
    put(bn,cell,t,HDR,fill=GREYFILL)
r=19
for p in order:
    put(bn,f'A{r}',p)
    put(bn,f'B{r}',round(pie[p]["Children"],4),BLACK,PCT); put(bn,f'C{r}',round(pie[p]["Adults"],4),BLACK,PCT)
    put(bn,f'D{r}',round(pie[p]["Disabled"],4),BLACK,PCT); put(bn,f'E{r}',round(pie[p]["Aged"],4),BLACK,PCT)
    put(bn,f'F{r}',f'=SUM(B{r}:E{r})',BLACK,PCT); put(bn,f'G{r}',round(final[p],2),BOLD,USD2); r+=1
put(bn,'A25',"Group $ (= service$ x share, summed)",BOLD)
for col,g in zip(['B','C','D','E'],G):
    put(bn,f'{col}25',f'=SUMPRODUCT($G$19:$G$24,{col}19:{col}24)',BOLD,USD2)
put(bn,'G25','=SUM(G19:G24)',BOLD,USD2)
put(bn,'A26',"Target group $ (column margins)",BLACK)
for col,g in zip(['B','C','D','E'],G):
    put(bn,f'{col}26',gt[g],BLACK,USD2)
put(bn,'A27',"Columns reconcile to within rounding; LTC reads ~90% Disabled + Aged, behavioral health ~79% Adults + Disabled.",BLACK,wrap=True)
bn.merge_cells('A27:G27'); bn['A27'].fill=YEL; bn.row_dimensions[27].height=26

# ---------- LEAKAGES ----------
lk=wb.create_sheet("Leakages")
for col,w in [('A',54),('B',16),('C',16),('D',42)]:
    lk.column_dimensions[col].width=w
put(lk,'A1',"Leakages - true losses off the flow",TITLE)
put(lk,'A2',"Plan-retained dollars and documented fraud are true leakages (gone from care). Per $100 of total Medicaid spending.",SUB)
section(lk,'A4',"1 - PLAN RETAINED (peels up off the two capitated lanes)",['A','B','C','D'])
for cell,t in [('A5','Component'),('B5','MCO lane'),('C5','Dual lane'),('D5','Note')]:
    put(lk,cell,t,HDR,fill=GREYFILL)
put(lk,'A6',"Capitation base (per $100 total)"); put(lk,'B6',"='Flow model'!D24",BLACK,USD2); put(lk,'C6',"='Flow model'!D25",BLACK,USD2)
put(lk,'A7',"Medical loss ratio (care delivered)"); put(lk,'B7',0.89,BLUE,PCT); put(lk,'C7',0.89,BLUE,PCT); put(lk,'D7',"Medicaid HBR ~88-91%",BLACK,wrap=True)
put(lk,'A8',"Plan admin (SG&A)"); put(lk,'B8',0.09,BLUE,PCT); put(lk,'C8',0.09,BLUE,PCT); put(lk,'D8',"Big-Five SG&A ~8-9%",BLACK,wrap=True)
put(lk,'A9',"Underwriting margin (1 - MLR - admin)"); put(lk,'B9','=1-B7-B8',BLACK,PCT); put(lk,'C9','=1-C7-C8',BLACK,PCT); put(lk,'D9',"thin; MLR floor 85%",BLACK,wrap=True)
put(lk,'A10',"  Plan admin retained"); put(lk,'B10','=B6*B8',BLACK,USD2); put(lk,'C10','=C6*C8',BLACK,USD2)
put(lk,'A11',"  Margin retained"); put(lk,'B11','=B6*B9',BLACK,USD2); put(lk,'C11','=C6*C9',BLACK,USD2)
put(lk,'A12',"     of which public-company earnings - ESTIMATE",BLACK,fill=YEL); put(lk,'B12',0.76,BLACK,USD2,fill=YEL); put(lk,'C12',"both lanes",BLACK,fill=YEL); put(lk,'D12',"subset of margin; pending 10-K (A5)",BLACK,fill=YEL,wrap=True)
put(lk,'A13',"Total plan-retained (both lanes)"); put(lk,'B13','=B10+B11+C10+C11',BOLD,USD2); put(lk,'D13',"about $5.61 per $100",BLACK,wrap=True)
put(lk,'A14',"  Plan administration (ex-earnings, both lanes)"); put(lk,'B14','=B13-B12',BLACK,USD2)
put(lk,'A15',"    of which by lane (retention share)"); put(lk,'B15','=B14*(B10+B11)/B13',BLACK,USD2); put(lk,'C15','=B14*(C10+C11)/B13',BLACK,USD2); put(lk,'D15',"MCO lane ~$3.81  |  dual-MCO lane ~$1.04",BLACK,wrap=True)
put(lk,'A16',"Dual plans carry the same carve as MCOs (A9); the dual lane's plan administration (~$1.04) is shown explicitly above. Duals' FFS spending has no carve and sits in the FFS lane.",BLACK,wrap=True); lk.merge_cells('A16:D16'); lk.row_dimensions[16].height=30
section(lk,'A17',"2 - DOCUMENTED FRAUD (true loss, off providers)",['A','B','C','D'])
put(lk,'A18',"MFCU recoveries, FY2024 ($M)"); put(lk,'B18',1400,BLUE,USD); put(lk,'D18',"OIG FY2024 ($961M crim + $407M civ)",BLACK,wrap=True)
put(lk,'A19',"As % of benefit spending"); put(lk,'B19',"=B18/'Flow model'!B6",BLACK,PCT2)
put(lk,'A20',"Per $100 total"); put(lk,'B20',"=B18/'Flow model'!B5*100",BLACK,USD2)
put(lk,'A21',"Recoveries understate total fraud; used as a documented true-loss proxy (A7). Drawn not-to-scale on the graphic.",BLACK,wrap=True); lk.merge_cells('A21:D21'); lk.row_dimensions[21].height=26
put(lk,'A23',"PERM improper-payment overlays were removed in v3 (flagged-not-lost; they confused the loss narrative).",BLACK,wrap=True); lk.merge_cells('A23:D23'); lk['A23'].fill=YEL

# ---------- SOURCES ----------
sc=wb.create_sheet("Sources")
for col,w in [('A',6),('B',52),('C',14),('D',60)]:
    sc.column_dimensions[col].width=w
put(sc,'B1',"Sources",TITLE); put(sc,'B2',"All accessed 8 June 2026. Federal fiscal year unless noted.",SUB)
for cell,t in [('A4','#'),('B4','Source'),('C4','Vintage'),('D4','Reference / URL')]:
    put(sc,cell,t,HDR,fill=GREYFILL)
srcs=[
 ("MACStats Feb 2026 - Ex.17 Total Benefit Spending by Category","FY2024","macpac.gov MACStats Ex.17"),
 ("MACStats Feb 2026 - Ex.16 Spending by Source of Funds (federal 64.7%)","FY2024","macpac.gov/publication/medicaid-spending-by-state-category-and-source-of-funds/"),
 ("MACStats Feb 2026 - Ex.21 Spending by Eligibility Group & Dual Status","FY2023","macpac.gov MACStats Ex.21"),
 ("CMS NHE Highlights - Medicaid $931.7B, 84.3M enrollees","CY2024","cms.gov/files/document/highlights.pdf"),
 ("HMA, 'New Insights on Medicaid Spending' - T-MSIS service mix (MC key)","CY2021/23","healthmanagement.com/.../New-Insights-on-Medicaid-Spending-1.pdf"),
 ("MedPAC-MACPAC, Duals Data Book - dual Medicaid $197.4B; Ex.12/16/18","CY2022","medpac.gov/.../Dec25_MedPAC_MACPAC_DualsDataBook-WEB508-FINAL.pdf"),
 ("HHS-OIG / MFCU Annual Report - $1.4B recovered","FY2024","oig.hhs.gov MFCU FY2024 statistical report"),
 ("Georgetown CCF - Big Five Medicaid MLR (Centene 90.9%, Molina 89.7%)","2024","ccf.georgetown.edu/2024/05/21/...big-five..."),
 ("Centene Corp SEC filings - Medicaid HBR & SG&A ratio","FY2024","sec.gov EDGAR CIK 0001071739"),
 ("Commonwealth Fund - Medicaid's Role in Mental Health & Substance Use Care ($58B MH + $17B SUD)","2019 data","commonwealthfund.org/publications/explainer/2025/may/medicaids-role-mental-health-and-substance-use-care"),
 ("Psychiatric Services (APA) - Medicaid behavioral health = 9.3-13% of spending","review","psychiatryonline.org/doi/full/10.1176/appi.ps.54.2.188"),
 ("KFF - Medicaid Mental Health & Substance Use: Expansion Trends","2025","kff.org/medicaid/medicaid-mental-health-and-substance-use-expansion-trends-and-the-fiscal-pressure-ahead/"),
]
r=5
for i,(s,v,u) in enumerate(srcs,1):
    put(sc,f'A{r}',i); put(sc,f'B{r}',s,BLACK,wrap=True); put(sc,f'C{r}',v); put(sc,f'D{r}',u,BLACK,wrap=True); sc.row_dimensions[r].height=28; r+=1

# ---------- ASSUMPTIONS ----------
asm=wb.create_sheet("Assumptions")
for col,w in [('A',6),('B',30),('C',82)]:
    asm.column_dimensions[col].width=w
put(asm,'B1',"Assumptions & Mapping Decisions",TITLE)
put(asm,'B2',"Each item is a modeling choice or known data gap. Yellow = needs attention / pending data.",SUB)
for cell,t in [('A4','#'),('B4','Topic'),('C4','Decision / gap')]:
    put(asm,cell,t,HDR,fill=GREYFILL)
asms=[
 ("Basis: the $100 combined dollar","The unit is $100 of TOTAL Medicaid outlays (benefits + admin), split federal $64.70 / state $35.30 at the 64.7% benefit share. Admin is jointly funded, so the true federal share of the combined dollar is ~64%; the headline 64.7% is used for the entry split.",True),
 ("Managed-care service allocation","Capitation is one CMS-64 line. The two modeled lanes are itemised across nodes with the HMA T-MSIS service mix, which prices MC encounters at FFS/Medicare rates. HMA 'Professional' unbundles hospital-based physician work that CMS-64 bundles into Hospital, so Physicians & clinics is overstated and Hospitals understated vs CMS-64 convention.",True),
 ("Eligibility-group vintage","Group spending shares are FY2023 (latest T-MSIS) applied to FY2024 totals.",True),
 ("FQHC handling (MERGED)","FQHC is clean (~$19B) on the FFS side but not separable in the HMA mix on the capitated side. Rather than show an FFS-only FQHC node that understates it, FQHC is MERGED into Physicians & clinics - completing the clinician number.",True),
 ("Public-company earnings carve","Shareholder net earnings (Medicaid-attributable) is a subset of margin, estimated at $0.76 per $100, pending 10-K segment allocation. Highest-uncertainty figure; never additive to margin.",True),
 ("Drugs net of rebates","Drugs shown net of manufacturer rebates (CMS-64 already net); rebate loop-back excluded.",False),
 ("Fraud proxy","Documented fraud uses MFCU recoveries, which understate total fraud. True loss off providers; drawn not-to-scale.",False),
 ("PERM removed","PERM improper-payment overlays were dropped in v3: they are flagged-not-lost and confused the loss narrative.",False),
 ("Duals scope & lane","Duals carry Medicaid dollars only. The dual LANE is dual managed-care capitation only (it carries the plan carve); duals' FFS spending sits in the FFS lane. Medicare premiums for duals ($27.8B) peel out at the agency.",False),
 ("Reporting-window mismatch","CMS-64 (federal FY), NHE (CY), MCO 10-Ks (company FY), Duals Data Book (CY2022), HMA mix (CY2021) differ. Anchored to FY2024 where possible; others are cross-checks.",True),
 ("Dual capitation split (Lane 2)","Dual MC capitation triangulated at ~$106B (band $100-115B) from the CY2022 Duals Data Book. The model's softest number.",True),
 ("Behavioral health (cross-cutting carve)","BH is ~9.3-13% of Medicaid spending (APA/Psychiatric Services claims-based review) and Medicaid is the largest U.S. BH payer (~24% of all national BH spending, MACPAC). BH has NO single CMS-64 line - it sits in inpatient psych (Hospitals), psychiatry/therapy/CMHC (Physicians & clinics), rehab/residential/community/case-management (Other, its largest source), psychotropics (Rx) and a little psychiatric residential (LTC). The $9.51 node (~11% of service dollars) is carved proportionally from those five lines; per-line splits are estimates pending T-MSIS. Its beneficiary pie (heavy Adults + Disabled) is likewise estimated.",True),
]
r=5
for i,(t,d,flag) in enumerate(asms,1):
    put(asm,f'A{r}',f"A{i}"); put(asm,f'B{r}',t,BOLD,wrap=True); put(asm,f'C{r}',d,BLACK,wrap=True)
    if flag:
        for c in ['A','B','C']: asm[f'{c}{r}'].fill=YEL
    asm.row_dimensions[r].height=58; r+=1

wb.save("/home/claude/medicaid_dollar_data.xlsx")
print("saved xlsx; provider total", round(sum(final.values()),2), "| nodes", {k:final[k] for k in order})
